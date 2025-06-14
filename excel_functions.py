import openpyxl
import config
from datetime import datetime


def create_excel_sheet(user_id: str):
    """
    Creates a new sheet in 'multiple_trading_sheet.xlsx' with the name from config.excel_sheet.
    Adds:
      - Table 1: ["Symbol", "Lots", "Last Price"] starting at A1
      - Leaves one empty column (D)
      - Table 2: ["Date", "Buy Trades", "Sell Trades", "Base Change"] starting at E1
      - Leaves one empty column (I)
      - Column header "Access Token" at J1
    If headers are already present in the correct format, does not change anything or delete previous data.
    """
    sheet_name = user_id
    file_name = "Excel sheets/" + user_id + ".xlsx"

    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Remove default sheet if present and empty
    if 'Sheet' in wb.sheetnames and len(wb['Sheet']['A']) == 0:
        wb.remove(wb['Sheet'])

    # Create or get the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Check if headers are already present in the correct format
    table1_headers = ["Symbol", "Lots", "Last Price"]
    table2_headers = ["Date", "Buy Trades", "Sell Trades", "Base Change", "Approximate Profit"]
    access_token_header = "Access Token"

    headers_ok = (
        [ws.cell(row=1, column=i).value for i in range(1, 4)] == table1_headers and
        [ws.cell(row=1, column=i).value for i in range(5, 10)] == table2_headers and
        ws.cell(row=1, column=11).value == access_token_header
    )

    if not headers_ok:
        # Only update headers, do not delete any data below
        for col, header in enumerate(table1_headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for col, header in enumerate(table2_headers, start=5):
            ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=11, value=access_token_header)

    wb.save(file_name)


def get_lots_for_symbol(user_id, symbol):
    """
    Reads the 'Lots' value for the given symbol from the sheet in 'multiple_trading_sheet.xlsx'.
    """
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().lower() == symbol.lower():
            return row[1]
    return None

def get_last_price_for_symbol(user_id: str, symbol):
    """
    Reads the 'Last Price' value for the given symbol from the sheet in 'multiple_trading_sheet.xlsx'.
    """
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]).strip().lower() == symbol.lower():
            return row[2]
    return None

def upsert_symbol_row(user_id, symbol, lots, last_price):
    """
    Inserts a new row with symbol, lots, and last price into the first available row
    of the first table (columns A, B, C) in the sheet.
    If the symbol already exists in column A, updates the lots and last price in the same row.
    """
    create_excel_sheet(user_id)  # Ensure the sheet is created before upserting
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    symbol = str(symbol).strip()
    found = False

    # Only consider rows where column A, B, C are part of the first table (i.e., before any empty row or before table 2)
    for row in ws.iter_rows(min_row=2, max_col=3):
        cell_symbol = str(row[0].value).strip() if row[0].value else ""
        # Stop if we hit an empty row (first table ends)
        if not any(cell.value for cell in row):
            break
        if cell_symbol.lower() == symbol.lower():
            row[1].value = lots
            row[2].value = last_price
            found = True
            break

    if not found:
        # Find the first empty row in columns A, B, C (first table)
        insert_row = 2
        while True:
            if not any(ws.cell(row=insert_row, column=col).value for col in range(1, 4)):
                ws.cell(row=insert_row, column=1, value=symbol)
                ws.cell(row=insert_row, column=2, value=lots)
                ws.cell(row=insert_row, column=3, value=last_price)
                break
            insert_row += 1

    wb.save(file_name)

def delete_all_data(user_id: str):
    """
    Deletes all data from the sheet in 'multiple_trading_sheet.xlsx'.
    """
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    # Clear all rows except the header
    ws.delete_rows(2, ws.max_row)

    wb.save(file_name)
    print(f"All data cleared from '{sheet_name}' in '{file_name}'.")

def read_symbols(user_id: str):
    """
    Reads all the symbols from the sheet in 'multiple_trading_sheet.xlsx'.
    Returns a list of symbols.
    """
    create_excel_sheet(user_id)  # Ensure the sheet is created before reading
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    symbols = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Check if the symbol cell is not empty
            symbols.append(row[0].strip())
    
    return symbols


def append_trading_orders(user_id: str, buy_trades, sell_trades, base_change):
    """
    Appends a new trading order with the current date, buy trades, sell trades, and base change
    to the first empty row of the second table (columns E-H) in the sheet.
    """
    create_excel_sheet(user_id)  # Ensure the sheet is created before appending
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    # Find the first empty row in columns E-H (second table)
    row_num = 2
    while True:
        # If all columns E-H are empty, this is our row
        if not any(ws.cell(row=row_num, column=col).value for col in range(5, 9)):
            break
        row_num += 1

    # Write values in their respective columns: Date (E), Buy Trades (F), Sell Trades (G), Base Change (H)
    ws.cell(row=row_num, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(row=row_num, column=6, value=buy_trades)
    ws.cell(row=row_num, column=7, value=sell_trades)
    ws.cell(row=row_num, column=8, value=base_change)
    ws.cell(row=row_num, column=9, value=(base_change + sell_trades) * 40)

    wb.save(file_name)
    print(f"New trading orders appended to '{sheet_name}' in '{file_name}' at columns E-H (row {row_num}).")

def get_access_token(user_id):
    """
    Reads the access token from cell J1 in the sheet in 'multiple_trading_sheet.xlsx'.
    Returns the access token as a string.
    """
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    access_token = ws.cell(row=2, column=11).value
    return access_token if access_token else None

def set_access_token(user_id: str, token):
    """
    Sets the access token in cell J1 of the sheet in 'multiple_trading_sheet.xlsx'.
    If the cell already contains a value, it will be overwritten.
    """
    file_name = "Excel sheets/" + user_id + ".xlsx"
    sheet_name = user_id

    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]

    ws.cell(row=2, column=11, value=token)

    wb.save(file_name)
    print(f"Access token set in '{sheet_name}' of '{file_name}'.")
    
if __name__ == "__main__":
    # print(get_lots_for_symbol('nbcc'))
    print(get_last_price_for_symbol("PQU213", 'CGCL'))
    print(get_lots_for_symbol('PQU213', 'CGCL'))
    # upsert_symbol_row('MSUMI', 15, 130.00)
    # print(read_symbols())
    # upsert_symbol_row('ab', 20, 125.00)
    # create_excel_sheet()
    # append_trading_orders(6, 3, 3)
    # print(get_access_token())
